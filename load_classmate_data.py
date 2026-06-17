"""
匯入同學提供的 Excel 資料（資料.xlsx）到資料庫
執行方式：python load_classmate_data.py

注意：執行前建議先清掉 db.sqlite3 再 migrate，確保資料乾淨
"""
import os
import django
import openpyxl
from datetime import date, datetime, timedelta
from pathlib import Path

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'library.settings')
django.setup()

from django.contrib.auth.models import User
from libraryapp.models import UserProfile, Book, Copy, Loan, Reservation


# ============== 設定 ==============
_here = Path(__file__).parent
# 支援中文或英文檔名（哪個存在就用哪個）
if (_here / '資料.xlsx').exists():
    EXCEL_PATH = _here / '資料.xlsx'
else:
    EXCEL_PATH = _here / 'data.xlsx'

# Excel 狀態碼 → 系統字串
USER_STATE_MAP = {1: 'active', 0: 'suspended', '1': 'active', '0': 'suspended'}
COPY_STATUS_MAP = {1: 'available', 2: 'borrowed', 3: 'available'}  # 3 是預約保留中，視為可借（保留的書）
RESERVATION_STATUS_MAP = {1: 'waiting', 2: 'waiting', 3: 'fulfilled'}  # 2 已到館待取也算 waiting


def clean(value):
    """清理空白、tab、None"""
    if value is None:
        return ''
    return str(value).strip().replace('\t', '')


def to_date(value):
    """轉成 date 物件"""
    if value is None or value == 'NULL':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def to_isbn(value):
    """ISBN 轉成字串（Excel 會把長數字變成 float）"""
    if value is None:
        return ''
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip()


# ============== 讀取 Excel ==============
print('📂 讀取 Excel...')
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)


# ============== 1. 匯入 USER ==============
print('\n=== 1. 匯入使用者 ===')
ws = wb['USER']
user_id_map = {}  # Excel 的 User_id → Django 的 user 物件

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    user_id, user_name, email, password, phone, state = row[:6]

    # 帳號用「名字英文部分」或從 email 取，這裡簡單用 email 前綴
    username_from_email = clean(email).split('@')[0] if email else f'user{int(user_id)}'
    username = username_from_email

    # 處理重複帳號
    if User.objects.filter(username=username).exists():
        username = f'{username}_{int(user_id)}'

    user, created = User.objects.get_or_create(
        username=username,
        defaults={
            'email': clean(email),
            'first_name': clean(user_name),  # 中文名放 first_name
        }
    )
    if created:
        user.set_password(str(int(password)) if password else '1234')
        user.save()

    state_str = USER_STATE_MAP.get(int(state) if state is not None else 1, 'active')
    UserProfile.objects.update_or_create(
        user=user,
        defaults={'phone': clean(phone), 'current_state': state_str}
    )

    user_id_map[int(user_id)] = user
    state_label = '正常' if state_str == 'active' else '停權'
    print(f'  ✓ #{int(user_id)} {clean(user_name)} ({username}) [{state_label}]')

print(f'共匯入 {len(user_id_map)} 個使用者')


# ============== 2. 匯入 BOOK ==============
print('\n=== 2. 匯入書籍 ===')
ws = wb['BOOK']
book_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    isbn, title, author, publisher, pub_date, category, location = row[:7]

    isbn_str = to_isbn(isbn)
    title_str = clean(title) if title is not None else ''
    # 處理書名是純數字的情況（例如 1984）
    if isinstance(title, (int, float)) and not isinstance(title, bool):
        title_str = str(int(title))

    book, created = Book.objects.update_or_create(
        isbn=isbn_str,
        defaults={
            'title': title_str,
            'author': clean(author),
            'publisher': clean(publisher),
            'publish_date': to_date(pub_date),
            'category': clean(category),
            'location': clean(location),
        }
    )
    book_count += 1
    print(f'  ✓ {isbn_str} 《{title_str}》 - {clean(author)}')

print(f'共匯入 {book_count} 本書')


# ============== 3. 匯入 COPY ==============
print('\n=== 3. 匯入副本 ===')
ws = wb['COPY']
copy_id_map = {}  # Excel 的 Copy_id → Django 的 copy 物件
copy_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    copy_id, isbn, status = row[:3]

    isbn_str = to_isbn(isbn)
    try:
        book = Book.objects.get(isbn=isbn_str)
    except Book.DoesNotExist:
        print(f'  ✗ Copy #{int(copy_id)} 對應的書籍 ISBN {isbn_str} 不存在！跳過')
        continue

    status_str = COPY_STATUS_MAP.get(int(status) if status else 1, 'available')

    # 用 Django 自動 ID（不指定 id，避免衝突）
    copy = Copy.objects.create(book=book, status=status_str)
    copy_id_map[int(copy_id)] = copy
    copy_count += 1

print(f'共匯入 {copy_count} 個副本')


# ============== 4. 匯入 LOAN ==============
print('\n=== 4. 匯入借閱記錄 ===')
# LOAN 表中有些 Return_date 寫成 2024 是錯的（應該是 2026），自動修正
ws = wb['LOAN']
loan_count = 0
TODAY = date.today()

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    # 跳過 RESERVATION 區塊（同 sheet 內混入的）
    if isinstance(row[0], str) and 'RESERVATION' in str(row[0]).upper():
        continue

    loan_id, user_id, copy_id, isbn, borrow_date, due_date, return_date = row[:7]

    user = user_id_map.get(int(user_id))
    copy = copy_id_map.get(int(copy_id))
    isbn_str = to_isbn(isbn)
    try:
        book = Book.objects.get(isbn=isbn_str)
    except Book.DoesNotExist:
        print(f'  ✗ Loan #{int(loan_id)} 對應的書籍不存在')
        continue

    if not user or not copy:
        print(f'  ✗ Loan #{int(loan_id)} 對應的 user 或 copy 不存在')
        continue

    borrow = to_date(borrow_date)
    due = to_date(due_date)
    ret = to_date(return_date)

    # 修正：如果 Return_date 是 2024 但 Borrow_date 是 2026（明顯打錯），改成 borrow + 幾天
    if ret and borrow and ret.year < borrow.year:
        ret = borrow + timedelta(days=10)
        fix_note = ' (Return_date 已修正)'
    else:
        fix_note = ''

    # 如果沒有 due_date 就用 borrow_date + 14 天
    if not due and borrow:
        due = borrow + timedelta(days=14)

    Loan.objects.filter(id=Loan.objects.create(
        user=user,
        copy=copy,
        book=book,
        due_date=due if due else TODAY + timedelta(days=14),
        return_date=ret,
    ).id).update(borrow_date=borrow if borrow else TODAY)
    loan_count += 1
    status = f'已還 {ret}' if ret else f'借閱中 (應還 {due})'
    print(f'  ✓ Loan #{int(loan_id)} {user.first_name} 借《{book.title[:15]}》— {status}{fix_note}')

print(f'共匯入 {loan_count} 筆借閱記錄')


# ============== 5. 匯入 RESERVATION ==============
print('\n=== 5. 匯入預約記錄 ===')
ws = wb['RESERVATION']
res_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue

    res_id, user_id, isbn, res_date, status = row[:5]

    user = user_id_map.get(int(user_id))
    isbn_str = to_isbn(isbn)
    try:
        book = Book.objects.get(isbn=isbn_str)
    except Book.DoesNotExist:
        print(f'  ✗ Reservation #{int(res_id)} 對應的書籍不存在')
        continue

    if not user:
        print(f'  ✗ Reservation #{int(res_id)} 對應的 user 不存在')
        continue

    status_str = RESERVATION_STATUS_MAP.get(int(status) if status else 1, 'waiting')

    res = Reservation.objects.create(
        user=user,
        book=book,
        status=status_str,
    )
    # 繞過 auto_now_add，設定原始 reservation_date
    Reservation.objects.filter(id=res.id).update(reservation_date=to_date(res_date) or TODAY)
    res_count += 1
    status_label = {'waiting': '等待中', 'fulfilled': '已取書', 'cancelled': '已取消'}[status_str]
    print(f'  ✓ Reservation #{int(res_id)} {user.first_name} 預約《{book.title[:15]}》[{status_label}]')

print(f'共匯入 {res_count} 筆預約記錄')


# ============== 6. 建立一個 admin 帳號（方便後台管理） ==============
print('\n=== 6. 建立管理員 ===')
if not User.objects.filter(username='admin').exists():
    admin = User.objects.create_superuser(username='admin', password='admin123', email='admin@library.local')
    UserProfile.objects.create(user=admin, phone='', current_state='active')
    print('  ✓ admin / admin123（Django 後台用）')
else:
    print('  - admin 已存在')


# ============== 完成 ==============
print('\n' + '='*50)
print('🎉 全部匯入完成！')
print('='*50)
print(f'\n📊 統計：')
print(f'  使用者：{User.objects.count()} 個')
print(f'  書籍：  {Book.objects.count()} 本')
print(f'  副本：  {Copy.objects.count()} 個')
print(f'  借閱：  {Loan.objects.count()} 筆 (進行中 {Loan.objects.filter(return_date__isnull=True).count()})')
print(f'  預約：  {Reservation.objects.count()} 筆')
print(f'\n🔑 可用帳號（密碼都是 1234，管理員是 admin123）：')
print(f'  讀者：daming / hsiaohua / jianguo / yating(停權) / elvan / david')
print(f'        beauty / mary / goan / john')
print(f'  管理員：admin / admin123')
